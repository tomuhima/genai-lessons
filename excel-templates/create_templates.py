"""
経営管理Excelテンプレート作成スクリプト
月次テンプレート + 年間サマリーテンプレート
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ============================================================
# マスターデータ
# ============================================================

# (名前, 種別, インボイス登録)
VENDORS = [
    ("梶原通信",               "個人", "あり"),
    ("秀電工",                 "個人", "あり"),
    ("井本貴史",               "個人", "なし"),
    ("LLS電気",                "個人", "あり"),
    ("株式会社トラストテクノス", "法人", "あり"),
    ("株式会社RISE",           "法人", "あり"),
]

# (名前, 締日, 入金サイト)
CLIENTS_MASTER = [
    ("港振興業株式会社",              "月末",   "翌月末"),
    ("株式会社トラストテクノス",       "月末",   "翌月末"),
    ("有限会社平成システム",           "20日",   "翌月末"),
    ("FGE合同会社",                   "月末",   "翌月末"),
    ("mtr株式会社",                   "月末",   "翌月末"),
    ("株式会社オークコミュニケーション", "月末",  "翌々月10日"),
    ("株式会社ライズ",                "確認中",  "確認中"),
    ("千里スカイハイツ管理組合",       "月末",   "翌々月末"),
    ("菊次",                          "月末",   "翌月末"),
    ("ページ",                        "都度",   "都度"),
]
CLIENTS = [c[0] for c in CLIENTS_MASTER]

# 人件費シート対象（代表者・外注は除く）
EMPLOYEES = ["丸田翔吾", "源地健史"]

EXPENSE_CATEGORIES = [
    "材料費", "燃料費", "高速料金", "駐車場代",
    "工具・消耗品費", "通信費", "交際費", "会議費",
    "事務用品費", "広告宣伝費", "研修費", "福利厚生費",
    "車両維持費", "地代家賃", "保険料", "雑費",
]

PAYMENT_METHODS = ["現金", "アメックス", "イオンカード"]

MONTHS = ["1月", "2月", "3月", "4月", "5月", "6月",
          "7月", "8月", "9月", "10月", "11月", "12月"]

# 燃料費単価（円/km）
FUEL_RATES = {
    "野添優":                   20,
    "丸田翔吾":                 20,
    "源地健史":                 20,
    "服部秀一":                 10,
    "井本貴史":                 10,
    "梶原通信":                 20,
    "LLS電気":                  20,
    "株式会社トラストテクノス":  20,
    "株式会社RISE":             20,
}

# 複数人対応（人員列を追加）
MULTI_PERSON_NAMES = {"梶原通信", "LLS電気", "株式会社トラストテクノス", "株式会社RISE"}

SHIFT_TYPES = ["昼勤", "夜勤", "半日（午前）", "半日（午後）"]

# ============================================================
# スタイル定義
# ============================================================
def header_style(color="1F4E79"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=10)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return fill, font, alignment

def subheader_style(color="2E75B6"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=10)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return fill, font, alignment

def border_thin():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def apply_headers(ws, headers, row=1, color="1F4E79"):
    fill, font, alignment = header_style(color)
    b = border_thin()
    for col, (text, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = b
        set_col_width(ws, col, width)
    ws.row_dimensions[row].height = 30

def data_row_style(ws, row, n_cols, alt=False):
    fill = PatternFill("solid", fgColor="EBF3FB" if alt else "FFFFFF")
    b = border_thin()
    alignment = Alignment(vertical="center")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = b
        cell.alignment = alignment

def add_dropdown(ws, col_letter, start_row, end_row, items):
    formula = '"' + ','.join(items) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)

def freeze_and_filter(ws, freeze_cell, filter_range):
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = filter_range


# ============================================================
# シート作成関数
# ============================================================

def create_expense_sheet(wb):
    """経費シート"""
    ws = wb.create_sheet("経費")
    ws.sheet_view.showGridLines = True

    headers = [
        ("No",          4),
        ("日付",        12),
        ("経費種類",    16),
        ("支払方法",    14),
        ("店名・支払先", 20),
        ("案件名",      20),
        ("金額（税込）", 14),
        ("備考",        24),
    ]
    apply_headers(ws, headers, row=1, color="1F4E79")

    DATA_ROWS = 200
    for i in range(DATA_ROWS):
        r = i + 2
        alt = (i % 2 == 1)
        data_row_style(ws, r, len(headers), alt)
        # No列（自動）
        ws.cell(r, 1).value = f'=IF(B{r}<>"",ROW()-1,"")'
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
        # 日付列フォーマット
        ws.cell(r, 2).number_format = "YYYY/MM/DD"
        # 金額列フォーマット
        ws.cell(r, 7).number_format = '#,##0'

    # ドロップダウン
    add_dropdown(ws, "C", 2, DATA_ROWS + 1, EXPENSE_CATEGORIES)
    add_dropdown(ws, "D", 2, DATA_ROWS + 1, PAYMENT_METHODS)

    freeze_and_filter(ws, "B2", f"A1:H{DATA_ROWS + 1}")

    # 合計行
    sum_row = DATA_ROWS + 2
    ws.cell(sum_row, 6).value = "合計"
    ws.cell(sum_row, 6).font = Font(bold=True)
    ws.cell(sum_row, 7).value = f"=SUM(G2:G{DATA_ROWS + 1})"
    ws.cell(sum_row, 7).number_format = '#,##0'
    ws.cell(sum_row, 7).font = Font(bold=True)


def create_subcontractor_sheet(wb):
    """外注管理シート"""
    ws = wb.create_sheet("外注管理")

    headers = [
        ("No",           4),
        ("登録日",       12),
        ("業者名",       22),
        ("種別",         8),
        ("インボイス",   10),
        ("案件名",       20),
        ("請求額(税抜)", 14),
        ("消費税(10%)",  13),
        ("合計(税込)",   14),
        ("請求日",       12),
        ("支払期限",     12),
        ("支払日",       12),
        ("支払額",       13),
        ("状況",         10),
    ]
    apply_headers(ws, headers, row=1, color="375623")

    DATA_ROWS = 200
    vendor_names = [v[0] for v in VENDORS]

    # インボイスなし警告色
    orange_fill = PatternFill("solid", fgColor="FFE4B5")

    for i in range(DATA_ROWS):
        r = i + 2
        alt = (i % 2 == 1)
        data_row_style(ws, r, len(headers), alt)
        ws.cell(r, 1).value = f'=IF(C{r}<>"",ROW()-1,"")'
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
        # 種別：マスターから自動取得（A列=業者名, B列=種別）
        ws.cell(r, 4).value = f'=IF(C{r}="","",IFERROR(VLOOKUP(C{r},マスター!A:B,2,0),""))'
        ws.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
        # インボイス：マスターから自動取得（A列=業者名, C列=インボイス）
        ws.cell(r, 5).value = f'=IF(C{r}="","",IFERROR(VLOOKUP(C{r},マスター!A:C,3,0),""))'
        ws.cell(r, 5).alignment = Alignment(horizontal="center", vertical="center")
        # 消費税自動計算
        ws.cell(r, 8).value = f'=IF(G{r}="","",G{r}*0.1)'
        ws.cell(r, 8).number_format = '#,##0'
        # 合計自動計算
        ws.cell(r, 9).value = f'=IF(G{r}="","",G{r}+H{r})'
        ws.cell(r, 9).number_format = '#,##0'
        # 支払期限：月末締め翌月末
        ws.cell(r, 11).value = f'=IF(J{r}="","",EOMONTH(J{r},1))'
        ws.cell(r, 11).number_format = "YYYY/MM/DD"
        # 状況：支払日があれば支払済
        ws.cell(r, 14).value = f'=IF(C{r}="","",IF(L{r}<>"","支払済","未払"))'
        ws.cell(r, 14).alignment = Alignment(horizontal="center", vertical="center")

        for col in [2, 10, 12]:
            ws.cell(r, col).number_format = "YYYY/MM/DD"
        for col in [7, 13]:
            ws.cell(r, col).number_format = '#,##0'

    add_dropdown(ws, "C", 2, DATA_ROWS + 1, vendor_names)
    freeze_and_filter(ws, "B2", f"A1:N{DATA_ROWS + 1}")

    # 条件付き書式（未払は赤・支払済は緑）
    red_fill = PatternFill("solid", fgColor="FFE0E0")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    ws.conditional_formatting.add(
        f"N2:N{DATA_ROWS + 1}",
        FormulaRule(formula=['N2="未払"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"N2:N{DATA_ROWS + 1}",
        FormulaRule(formula=['N2="支払済"'], fill=green_fill)
    )
    # インボイスなしは橙色で警告
    ws.conditional_formatting.add(
        f"E2:E{DATA_ROWS + 1}",
        FormulaRule(formula=['E2="なし"'], fill=orange_fill)
    )

    # 合計行（列番号はヘッダー追加後: G=7,H=8,I=9,M=13,N=14）
    sum_row = DATA_ROWS + 2
    ws.cell(sum_row, 6).value = "合計"
    ws.cell(sum_row, 6).font = Font(bold=True)
    for col in [7, 8, 9, 13]:  # 請求額,消費税,合計,支払額
        ws.cell(sum_row, col).value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{DATA_ROWS+1})"
        ws.cell(sum_row, col).number_format = '#,##0'
        ws.cell(sum_row, col).font = Font(bold=True)
    # 未払合計
    ws.cell(sum_row, 14).value = f'=COUNTIF(N2:N{DATA_ROWS+1},"未払")&"件未払"'
    ws.cell(sum_row, 14).font = Font(bold=True, color="CC0000")


def create_sales_sheet(wb):
    """売上管理シート"""
    ws = wb.create_sheet("売上管理")

    headers = [
        ("No",          4),
        ("登録日",      12),
        ("得意先名",    24),
        ("案件名",      22),
        ("請求額(税抜)", 14),
        ("消費税(10%)", 13),
        ("合計(税込)",  14),
        ("請求日",      12),
        ("入金期限",    12),
        ("入金日",      12),
        ("入金額",      13),
        ("状況",        12),
    ]
    apply_headers(ws, headers, row=1, color="7B2C2C")

    DATA_ROWS = 200
    for i in range(DATA_ROWS):
        r = i + 2
        alt = (i % 2 == 1)
        data_row_style(ws, r, len(headers), alt)
        ws.cell(r, 1).value = f'=IF(C{r}<>"",ROW()-1,"")'
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 6).value = f'=IF(E{r}="","",E{r}*0.1)'
        ws.cell(r, 6).number_format = '#,##0'
        ws.cell(r, 7).value = f'=IF(E{r}="","",E{r}+F{r})'
        ws.cell(r, 7).number_format = '#,##0'
        # 入金状況の判定
        ws.cell(r, 12).value = (
            f'=IF(C{r}="","",IF(J{r}="","未入金",'
            f'IF(K{r}>=G{r},"入金済","一部入金")))'
        )
        ws.cell(r, 12).alignment = Alignment(horizontal="center", vertical="center")

        for col in [2, 8, 9, 10]:
            ws.cell(r, col).number_format = "YYYY/MM/DD"
        for col in [5, 11]:
            ws.cell(r, col).number_format = '#,##0'

    add_dropdown(ws, "C", 2, DATA_ROWS + 1, CLIENTS)

    # 状況ドロップダウンは数式で自動だが手動上書きも可能に
    freeze_and_filter(ws, "B2", f"A1:L{DATA_ROWS + 1}")

    # 条件付き書式
    red_fill = PatternFill("solid", fgColor="FFE0E0")
    yellow_fill = PatternFill("solid", fgColor="FFFACD")
    green_fill = PatternFill("solid", fgColor="E2EFDA")
    ws.conditional_formatting.add(
        f"L2:L{DATA_ROWS + 1}",
        FormulaRule(formula=['L2="未入金"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"L2:L{DATA_ROWS + 1}",
        FormulaRule(formula=['L2="一部入金"'], fill=yellow_fill)
    )
    ws.conditional_formatting.add(
        f"L2:L{DATA_ROWS + 1}",
        FormulaRule(formula=['L2="入金済"'], fill=green_fill)
    )

    # 合計行
    sum_row = DATA_ROWS + 2
    ws.cell(sum_row, 4).value = "合計"
    ws.cell(sum_row, 4).font = Font(bold=True)
    for col in [5, 6, 7, 11]:
        ws.cell(sum_row, col).value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{DATA_ROWS+1})"
        ws.cell(sum_row, col).number_format = '#,##0'
        ws.cell(sum_row, col).font = Font(bold=True)
    ws.cell(sum_row, 12).value = f'=COUNTIF(L2:L{DATA_ROWS+1},"未入金")&"件未入金"'
    ws.cell(sum_row, 12).font = Font(bold=True, color="CC0000")


def create_payroll_sheet(wb):
    """人件費シート"""
    ws = wb.create_sheet("人件費")

    headers = [
        ("社員名",   16),
        ("支給総額", 14),
        ("控除合計", 14),
        ("手取り",   14),
        ("支払日",   12),
        ("備考",     24),
    ]
    apply_headers(ws, headers, row=1, color="4B4B8F")

    # 社員行を最初から用意
    all_employees = EMPLOYEES + [""] * 8  # 最大13名まで
    for i, emp in enumerate(all_employees[:13]):
        r = i + 2
        alt = (i % 2 == 1)
        data_row_style(ws, r, len(headers), alt)
        if emp:
            ws.cell(r, 1).value = emp
        for col in [2, 3, 4]:
            ws.cell(r, col).number_format = '#,##0'
        ws.cell(r, 5).number_format = "YYYY/MM/DD"
        ws.cell(r, 4).value = f'=IF(B{r}="","",B{r}-C{r})'

    # 合計行
    sum_row = 15
    ws.cell(sum_row, 1).value = "合計"
    ws.cell(sum_row, 1).font = Font(bold=True)
    for col in [2, 3, 4]:
        ws.cell(sum_row, col).value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{sum_row - 1})"
        ws.cell(sum_row, col).number_format = '#,##0'
        ws.cell(sum_row, col).font = Font(bold=True)

    ws.freeze_panes = "B2"


def create_summary_sheet(wb):
    """月次サマリーシート"""
    ws = wb.create_sheet("月次サマリー")

    # タイトル
    ws.merge_cells("A1:D1")
    title = ws.cell(1, 1, "月次サマリー")
    title.font = Font(bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    fill_blue = PatternFill("solid", fgColor="D9E8F5")
    fill_green = PatternFill("solid", fgColor="E2EFDA")
    fill_red = PatternFill("solid", fgColor="FFE0E0")
    fill_orange = PatternFill("solid", fgColor="FFF2CC")
    fill_purple = PatternFill("solid", fgColor="EAE8F0")
    bold = Font(bold=True)
    b = border_thin()

    def set_label(r, c, text, fill, font=None):
        cell = ws.cell(r, c, text)
        cell.fill = fill
        cell.border = b
        cell.font = font or Font(bold=True)
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22

    def set_value(r, c, formula, fmt='#,##0'):
        cell = ws.cell(r, c, formula)
        cell.number_format = fmt
        cell.border = b
        cell.alignment = Alignment(horizontal="right", vertical="center")

    DATA_ROWS = 200

    # ---- 売上 ----
    set_label(3, 1, "■ 売上", fill_blue)
    ws.merge_cells("A3:D3")

    set_label(4, 1, "請求済合計（税込）", fill_blue, bold)
    set_value(4, 2, f"=SUM(売上管理!G2:G{DATA_ROWS+1})")
    set_label(5, 1, "入金済合計", fill_blue, bold)
    set_value(5, 2, f"=SUMIF(売上管理!L2:L{DATA_ROWS+1},\"入金済\",売上管理!K2:K{DATA_ROWS+1})")
    set_label(6, 1, "未入金残高", fill_red, Font(bold=True, color="CC0000"))
    set_value(6, 2, f"=SUMIF(売上管理!L2:L{DATA_ROWS+1},\"未入金\",売上管理!G2:G{DATA_ROWS+1})")

    # ---- 外注費 ----
    set_label(8, 1, "■ 外注費", fill_green)
    ws.merge_cells("A8:D8")

    set_label(9, 1,  "請求受取合計（税込）", fill_green, bold)
    set_value(9, 2,  f"=SUM(外注管理!H2:H{DATA_ROWS+1})")
    set_label(10, 1, "支払済合計", fill_green, bold)
    set_value(10, 2, f"=SUMIF(外注管理!M2:M{DATA_ROWS+1},\"支払済\",外注管理!L2:L{DATA_ROWS+1})")
    set_label(11, 1, "未払残高", fill_red, Font(bold=True, color="CC0000"))
    set_value(11, 2, f"=SUMIF(外注管理!M2:M{DATA_ROWS+1},\"未払\",外注管理!H2:H{DATA_ROWS+1})")

    # ---- 経費 ----
    set_label(13, 1, "■ 経費", fill_orange)
    ws.merge_cells("A13:D13")

    for i, cat in enumerate(EXPENSE_CATEGORIES):
        r = 14 + i
        set_label(r, 1, cat, fill_orange, Font())
        set_value(r, 2, f'=SUMIF(経費!C2:C{DATA_ROWS+1},"{cat}",経費!G2:G{DATA_ROWS+1})')

    exp_sum_row = 14 + len(EXPENSE_CATEGORIES)
    set_label(exp_sum_row, 1, "経費合計", fill_orange, bold)
    set_value(exp_sum_row, 2, f"=SUM(経費!G2:G{DATA_ROWS+1})")

    # ---- 人件費 ----
    payroll_row = exp_sum_row + 2
    set_label(payroll_row, 1, "■ 人件費", fill_purple)
    ws.merge_cells(f"A{payroll_row}:D{payroll_row}")
    set_label(payroll_row + 1, 1, "支給総額合計", fill_purple, bold)
    set_value(payroll_row + 1, 2, "=SUM(人件費!B2:B14)")
    set_label(payroll_row + 2, 1, "手取り合計", fill_purple, bold)
    set_value(payroll_row + 2, 2, "=SUM(人件費!D2:D14)")

    # ---- 粗利 ----
    profit_row = payroll_row + 4
    ws.merge_cells(f"A{profit_row}:D{profit_row}")
    set_label(profit_row, 1, "■ 粗利（概算）", PatternFill("solid", fgColor="1F4E79"), Font(bold=True, color="FFFFFF", size=12))

    gross = f"=売上管理!G{DATA_ROWS+2}-外注管理!H{DATA_ROWS+2}-経費!G{DATA_ROWS+2}-人件費!B15"
    cell = ws.cell(profit_row + 1, 1, f"=SUM(売上管理!G2:G{DATA_ROWS+1})-SUM(外注管理!H2:H{DATA_ROWS+1})-SUM(経費!G2:G{DATA_ROWS+1})-SUM(人件費!B2:B14)")
    cell.number_format = '#,##0'
    cell.font = Font(bold=True, size=14)
    cell.border = b
    cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"A{profit_row+1}:D{profit_row+1}")
    ws.row_dimensions[profit_row + 1].height = 30

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


def create_master_sheet(wb):
    """マスターシート（非表示）"""
    ws = wb.create_sheet("マスター")

    # 業者マスター（A:C）
    ws.cell(1, 1, "業者名").font = Font(bold=True)
    ws.cell(1, 2, "種別").font = Font(bold=True)
    ws.cell(1, 3, "インボイス").font = Font(bold=True)
    for i, (name, vendor_type, invoice) in enumerate(VENDORS, 2):
        ws.cell(i, 1, name)
        ws.cell(i, 2, vendor_type)
        ws.cell(i, 3, invoice)

    # 得意先マスター（E:G列）
    ws.cell(1, 5, "得意先名").font = Font(bold=True)
    ws.cell(1, 6, "締日").font = Font(bold=True)
    ws.cell(1, 7, "入金サイト").font = Font(bold=True)
    for i, (name, closing, payment) in enumerate(CLIENTS_MASTER, 2):
        ws.cell(i, 5, name)
        ws.cell(i, 6, closing)
        ws.cell(i, 7, payment)

    # 経費種類マスター（I列）
    ws.cell(1, 9, "経費種類").font = Font(bold=True)
    for i, cat in enumerate(EXPENSE_CATEGORIES, 2):
        ws.cell(i, 9, cat)

    ws.sheet_state = "hidden"


# ============================================================
# 個人別稼働確認テンプレート作成
# ============================================================

def create_kado_template(name, fuel_rate, multi_person=False):
    """個人別稼働確認シート（複数現場・複数人対応）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "稼働確認"

    # 列定義（multi_personの場合は人員列を追加）
    if multi_person:
        headers = [
            ("No",        4),  ("日付",     12), ("区分",     10),
            ("会社名",   22),  ("物件名",   20), ("人員",     18),
            ("開始",      8),  ("終了",      8), ("休憩(分)",  8),
            ("実働時間", 10),  ("作業内容", 26),
            ("高速料金", 12),  ("移動距離", 10), ("燃料費",   10),
            ("駐車場",   10),  ("資材費",   10), ("経費合計", 12),
        ]
        C_CLIENT = 4; C_START = 7; C_END = 8; C_BREAK = 9
        C_HOURS  = 10; C_HW = 12; C_DIST = 13
        C_FUEL   = 14; C_PARK = 15; C_MAT = 16; C_TOTAL = 17
    else:
        headers = [
            ("No",        4),  ("日付",     12), ("区分",     10),
            ("会社名",   22),  ("物件名",   20),
            ("開始",      8),  ("終了",      8), ("休憩(分)",  8),
            ("実働時間", 10),  ("作業内容", 26),
            ("高速料金", 12),  ("移動距離", 10), ("燃料費",   10),
            ("駐車場",   10),  ("資材費",   10), ("経費合計", 12),
        ]
        C_CLIENT = 4; C_START = 6; C_END = 7; C_BREAK = 8
        C_HOURS  = 9; C_HW = 11; C_DIST = 12
        C_FUEL   = 13; C_PARK = 14; C_MAT = 15; C_TOTAL = 16

    n_cols = len(headers)
    sl = get_column_letter

    # タイトル行
    ws.merge_cells(f"A1:{sl(n_cols)}1")
    t = ws.cell(1, 1, f"稼働確認シート　{name}　（燃料費：{fuel_rate}円/km）")
    t.font      = Font(bold=True, size=12, color="FFFFFF")
    t.fill      = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ヘッダー行（row=2）
    apply_headers(ws, headers, row=2, color="2E75B6")

    # データ行（100行）
    DATA_ROWS = 100
    for i in range(DATA_ROWS):
        r   = i + 3
        alt = (i % 2 == 1)
        data_row_style(ws, r, n_cols, alt)

        cl = sl(C_CLIENT)

        # No（会社名があれば連番）
        ws.cell(r, 1).value     = f'=IF({cl}{r}<>"",ROW()-2,"")'
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")

        # 日付フォーマット
        ws.cell(r, 2).number_format = "MM/DD"

        # 開始・終了（時間フォーマット）
        ws.cell(r, C_START).number_format = "HH:MM"
        ws.cell(r, C_END).number_format   = "HH:MM"

        # 実働時間 = (終了-開始)*24 - 休憩分/60
        s_l = sl(C_START); e_l = sl(C_END); b_l = sl(C_BREAK)
        ws.cell(r, C_HOURS).value = (
            f'=IF({s_l}{r}="","",MAX(0,({e_l}{r}-{s_l}{r})*24-{b_l}{r}/60))'
        )
        ws.cell(r, C_HOURS).number_format = "0.0"
        ws.cell(r, C_HOURS).alignment     = Alignment(horizontal="center", vertical="center")

        # 燃料費 = 移動距離 × 単価
        d_l = sl(C_DIST)
        ws.cell(r, C_FUEL).value          = f'=IF({d_l}{r}="","",{d_l}{r}*{fuel_rate})'
        ws.cell(r, C_FUEL).number_format  = "#,##0"

        # 経費合計 = 高速+燃料費+駐車場+資材費
        hw_l = sl(C_HW); f_l = sl(C_FUEL); p_l = sl(C_PARK); m_l = sl(C_MAT)
        ws.cell(r, C_TOTAL).value         = (
            f'=IF({cl}{r}="","",SUM({hw_l}{r},{f_l}{r},{p_l}{r},{m_l}{r}))'
        )
        ws.cell(r, C_TOTAL).number_format = "#,##0"

        # 数値列フォーマット
        for col in [C_HW, C_DIST, C_PARK, C_MAT]:
            ws.cell(r, col).number_format = "#,##0"

    # ドロップダウン
    add_dropdown(ws, sl(3),       3, DATA_ROWS + 2, SHIFT_TYPES)  # 区分
    add_dropdown(ws, sl(C_CLIENT),3, DATA_ROWS + 2, CLIENTS)      # 会社名

    # 合計行
    sum_row = DATA_ROWS + 3
    ws.merge_cells(f"A{sum_row}:B{sum_row}")
    sc = ws.cell(sum_row, 1, "月次合計")
    sc.font      = Font(bold=True)
    sc.alignment = Alignment(horizontal="center", vertical="center")

    # 稼働日数
    ws.cell(sum_row, C_HOURS).value = f'=COUNTA(B3:B{sum_row-1})'
    ws.cell(sum_row, C_HOURS).font  = Font(bold=True)
    ws.cell(sum_row, C_HOURS).alignment = Alignment(horizontal="center", vertical="center")

    for col in [C_HW, C_FUEL, C_PARK, C_MAT, C_TOTAL]:
        c_l = sl(col)
        ws.cell(sum_row, col).value        = f'=SUM({c_l}3:{c_l}{sum_row-1})'
        ws.cell(sum_row, col).number_format = "#,##0"
        ws.cell(sum_row, col).font          = Font(bold=True)

    sum_fill = PatternFill("solid", fgColor="D9E8F5")
    b = border_thin()
    ws.row_dimensions[sum_row].height = 24
    for col in range(1, n_cols + 1):
        ws.cell(sum_row, col).fill   = sum_fill
        ws.cell(sum_row, col).border = b

    # フリーズ（タイトル+ヘッダー固定、No列固定）
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{sl(n_cols)}{sum_row - 1}"

    safe = name.replace("株式会社", "").replace("　", "").strip()
    path = os.path.join(OUTPUT_DIR, f"稼働確認_{safe}_テンプレート.xlsx")
    wb.save(path)
    print(f"  ✓ {path}")
    return path


# ============================================================
# 月次テンプレート作成
# ============================================================
def create_monthly_template():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    create_master_sheet(wb)
    create_summary_sheet(wb)
    create_expense_sheet(wb)
    create_subcontractor_sheet(wb)
    create_sales_sheet(wb)
    create_payroll_sheet(wb)

    # シート順を整える
    sheet_order = ["月次サマリー", "経費", "外注管理", "売上管理", "人件費", "マスター"]
    for i, name in enumerate(sheet_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    path = os.path.join(OUTPUT_DIR, "経営管理_テンプレート.xlsx")
    wb.save(path)
    print(f"✓ 月次テンプレート作成: {path}")
    return path


# ============================================================
# 年間サマリーテンプレート作成
# ============================================================
def create_annual_template():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- 月別収支シート ----
    ws1 = wb.create_sheet("月別収支")
    col_headers = [("月", 8), ("売上請求", 14), ("入金済", 14), ("未入金", 14),
                   ("外注費", 14), ("外注支払", 14), ("未払", 14),
                   ("経費", 14), ("人件費", 14), ("粗利（概算）", 16)]
    apply_headers(ws1, col_headers, row=1, color="1F4E79")
    for i, month in enumerate(MONTHS):
        r = i + 2
        alt = (i % 2 == 1)
        data_row_style(ws1, r, len(col_headers), alt)
        ws1.cell(r, 1, month).alignment = Alignment(horizontal="center", vertical="center")
        for col in range(2, len(col_headers) + 1):
            ws1.cell(r, col).number_format = '#,##0'
    # 合計行
    sum_row = 14
    ws1.cell(sum_row, 1, "合計").font = Font(bold=True)
    ws1.cell(sum_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    for col in range(2, len(col_headers) + 1):
        ws1.cell(sum_row, col).value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}13)"
        ws1.cell(sum_row, col).number_format = '#,##0'
        ws1.cell(sum_row, col).font = Font(bold=True)
    ws1.freeze_panes = "B2"

    # ---- 業者別外注費シート ----
    ws2 = wb.create_sheet("業者別外注費")
    vendor_headers = [("業者名", 24)] + [(m, 12) for m in MONTHS] + [("合計", 14)]
    apply_headers(ws2, vendor_headers, row=1, color="375623")
    for i, (name, *_) in enumerate(VENDORS):
        r = i + 2
        alt = (i % 2 == 1)
        data_row_style(ws2, r, len(vendor_headers), alt)
        ws2.cell(r, 1, name)
        for col in range(2, len(vendor_headers) + 1):
            ws2.cell(r, col).number_format = '#,##0'
        # 合計列
        ws2.cell(r, len(vendor_headers)).value = f"=SUM(B{r}:{get_column_letter(len(vendor_headers)-1)}{r})"
        ws2.cell(r, len(vendor_headers)).font = Font(bold=True)
    # 合計行
    vsum_row = len(VENDORS) + 2
    ws2.cell(vsum_row, 1, "合計").font = Font(bold=True)
    for col in range(2, len(vendor_headers) + 1):
        ws2.cell(vsum_row, col).value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{vsum_row-1})"
        ws2.cell(vsum_row, col).number_format = '#,##0'
        ws2.cell(vsum_row, col).font = Font(bold=True)
    ws2.freeze_panes = "B2"

    # ---- 得意先別売上シート ----
    ws3 = wb.create_sheet("得意先別売上")
    client_headers = [("得意先名", 28)] + [(m, 12) for m in MONTHS] + [("合計", 14)]
    apply_headers(ws3, client_headers, row=1, color="7B2C2C")
    for i, name in enumerate(CLIENTS):
        r = i + 2
        alt = (i % 2 == 1)
        data_row_style(ws3, r, len(client_headers), alt)
        ws3.cell(r, 1, name)
        for col in range(2, len(client_headers) + 1):
            ws3.cell(r, col).number_format = '#,##0'
        ws3.cell(r, len(client_headers)).value = f"=SUM(B{r}:{get_column_letter(len(client_headers)-1)}{r})"
        ws3.cell(r, len(client_headers)).font = Font(bold=True)
    # 合計行
    csum_row = len(CLIENTS) + 2
    ws3.cell(csum_row, 1, "合計").font = Font(bold=True)
    for col in range(2, len(client_headers) + 1):
        ws3.cell(csum_row, col).value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{csum_row-1})"
        ws3.cell(csum_row, col).number_format = '#,##0'
        ws3.cell(csum_row, col).font = Font(bold=True)
    ws3.freeze_panes = "B2"

    # ---- 経費種類別シート ----
    ws4 = wb.create_sheet("経費種類別")
    exp_headers = [("経費種類", 20)] + [(m, 12) for m in MONTHS] + [("合計", 14)]
    apply_headers(ws4, exp_headers, row=1, color="7B5C00")
    for i, cat in enumerate(EXPENSE_CATEGORIES):
        r = i + 2
        alt = (i % 2 == 1)
        data_row_style(ws4, r, len(exp_headers), alt)
        ws4.cell(r, 1, cat)
        for col in range(2, len(exp_headers) + 1):
            ws4.cell(r, col).number_format = '#,##0'
        ws4.cell(r, len(exp_headers)).value = f"=SUM(B{r}:{get_column_letter(len(exp_headers)-1)}{r})"
        ws4.cell(r, len(exp_headers)).font = Font(bold=True)
    esum_row = len(EXPENSE_CATEGORIES) + 2
    ws4.cell(esum_row, 1, "合計").font = Font(bold=True)
    for col in range(2, len(exp_headers) + 1):
        ws4.cell(esum_row, col).value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{esum_row-1})"
        ws4.cell(esum_row, col).number_format = '#,##0'
        ws4.cell(esum_row, col).font = Font(bold=True)
    ws4.freeze_panes = "B2"

    path = os.path.join(OUTPUT_DIR, "年間サマリー_テンプレート.xlsx")
    wb.save(path)
    print(f"✓ 年間サマリーテンプレート作成: {path}")
    return path


if __name__ == "__main__":
    create_monthly_template()
    create_annual_template()

    print("\n個人別稼働確認テンプレート作成中...")
    for person_name, rate in FUEL_RATES.items():
        multi = person_name in MULTI_PERSON_NAMES
        create_kado_template(person_name, rate, multi_person=multi)

    print("\n完了しました。")
