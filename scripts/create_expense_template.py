#!/usr/bin/env python3
"""
経費管理Excelテンプレート生成スクリプト
Usage: python3 create_expense_template.py [YYYY] [MM]
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

year = int(sys.argv[1]) if len(sys.argv) > 1 else 2026
month = int(sys.argv[2]) if len(sys.argv) > 2 else 4

filename = f"/tmp/経費管理_{year}年{month:02d}月.xlsx"

# カラー定義
C_BLUE_H   = "2E75B6"   # 売上ヘッダー
C_BLUE_BG  = "DDEEFF"   # 売上背景
C_GREEN_H  = "375623"   # 外注ヘッダー
C_GREEN_BG = "E2EFDA"   # 外注背景
C_ORG_H    = "C55A11"   # 経費ヘッダー
C_ORG_BG   = "FCE4D6"   # 経費背景
C_DARK_H   = "1F3864"   # 人件費・粗利ヘッダー
C_RED      = "FF0000"   # 未払残高・未入金

wb = openpyxl.Workbook()

# ==================== 月次サマリー ====================
ws = wb.active
ws.title = "月次サマリー"

def header(ws, row, label, bg, text_color="FFFFFF"):
    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(bold=True, color=text_color, size=11)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def data_row(ws, row, label, formula, bg=None, red=False):
    ca = ws[f"A{row}"]
    cb = ws[f"B{row}"]
    ca.value = label
    cb.value = formula
    ca.alignment = Alignment(vertical="center", indent=1)
    cb.alignment = Alignment(horizontal="right", vertical="center")
    cb.number_format = '#,##0'
    if red:
        ca.font = Font(color=C_RED)
    if bg:
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=bg)

# タイトル
ws.merge_cells("A1:E1")
ws["A1"].value = f"月次サマリー　{year}年{month:02d}月"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=C_DARK_H)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# ■ 売上
header(ws, 3, "■ 売上", C_BLUE_H)
data_row(ws, 4, "請求済合計（税込）", "=SUMIF('売上管理'!B:B,\"<>\",'売上管理'!G:G)", C_BLUE_BG)
data_row(ws, 5, "入金済合計",         "=SUMIF('売上管理'!L:L,\"入金済\",'売上管理'!K:K)", C_BLUE_BG)
data_row(ws, 6, "未入金残高",         "=B4-B5", C_BLUE_BG, red=True)

# ■ 外注費
header(ws, 8, "■ 外注費", C_GREEN_H)
data_row(ws, 9,  "請求受取合計（税込）", "=SUMIF('外注管理'!B:B,\"<>\",'外注管理'!I:I)", C_GREEN_BG)
data_row(ws, 10, "支払済合計",           "=SUMIF('外注管理'!N:N,\"支払済\",'外注管理'!M:M)", C_GREEN_BG)
data_row(ws, 11, "未払残高",             "=B9-B10", C_GREEN_BG, red=True)

# ■ 経費
header(ws, 13, "■ 経費", C_ORG_H)
expense_cats = [
    "材料費", "燃料費", "高速料金", "駐車場代", "工具・消耗品費",
    "通信費", "交際費", "会議費", "事務用品費", "広告宣伝費",
    "研修費", "福利厚生費", "車両維持費", "地代家賃", "保険料", "雑費"
]
for i, cat in enumerate(expense_cats):
    row = 14 + i
    data_row(ws, row, cat, f"=SUMIF('経費'!C:C,\"{cat}\",'経費'!G:G)")
data_row(ws, 30, "経費合計", "=SUM(B14:B29)")
ws["A30"].font = Font(bold=True)
ws["B30"].font = Font(bold=True)

# ■ 人件費
header(ws, 32, "■ 人件費", C_DARK_H)
data_row(ws, 33, "支給総額合計", "=SUM('人件費'!B2:B14)")
data_row(ws, 34, "手取り合計",   "=SUM('人件費'!D2:D14)")

# ■ 粗利（概算）
header(ws, 36, "■ 粗利（概算）", C_DARK_H)
ws.merge_cells("A37:D37")
ws["E37"].value = "=B5-B10-B30-B33"
ws["E37"].font = Font(bold=True, size=13, color="FFFFFF")
ws["E37"].fill = PatternFill("solid", fgColor=C_DARK_H)
ws["E37"].number_format = '#,##0'
ws["E37"].alignment = Alignment(horizontal="right", vertical="center")
ws.row_dimensions[37].height = 28

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 5
ws.column_dimensions["D"].width = 5
ws.column_dimensions["E"].width = 16
ws.sheet_view.showGridLines = False

# ==================== 経費シート ====================
ws_exp = wb.create_sheet("経費")
hdrs_exp = ["No.", "日付", "経費種類", "支払方法", "店名・支払先", "案件名", "金額（税込）", "備考"]
widths_exp = [6, 12, 16, 12, 20, 20, 14, 24]

for col, h in enumerate(hdrs_exp, 1):
    c = ws_exp.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_BLUE_H)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_exp.column_dimensions[get_column_letter(col)].width = widths_exp[col - 1]
ws_exp.row_dimensions[1].height = 24

for row in range(2, 201):
    bg = "EBF3FB" if row % 2 == 0 else "FFFFFF"
    for col in range(1, len(hdrs_exp) + 1):
        ws_exp.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
ws_exp.sheet_view.showGridLines = False

# ==================== 外注管理シート ====================
ws_sub = wb.create_sheet("外注管理")
hdrs_sub = ["No.", "登録日", "業者名", "種別", "インボイス", "案件名",
            "請求額（税抜）", "消費税（10%）", "合計（税込）",
            "請求日", "支払期限", "支払日", "支払額", "状況"]
widths_sub = [6, 12, 15, 10, 12, 20, 14, 14, 14, 12, 12, 12, 12, 10]

for col, h in enumerate(hdrs_sub, 1):
    c = ws_sub.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_GREEN_H)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sub.column_dimensions[get_column_letter(col)].width = widths_sub[col - 1]
ws_sub.row_dimensions[1].height = 24

for row in range(2, 201):
    bg = "E2EFDA" if row % 2 == 0 else "FFFFFF"
    for col in range(1, len(hdrs_sub) + 1):
        ws_sub.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
ws_sub.sheet_view.showGridLines = False

# ==================== 売上管理シート ====================
ws_sales = wb.create_sheet("売上管理")
hdrs_sales = ["No.", "登録日", "得意先名", "案件名",
              "請求額（税抜）", "消費税（10%）", "合計（税込）",
              "請求日", "入金期限", "入金日", "入金額", "状況"]
widths_sales = [6, 12, 16, 22, 14, 14, 14, 12, 12, 12, 12, 10]

for col, h in enumerate(hdrs_sales, 1):
    c = ws_sales.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="843C0C")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_sales.column_dimensions[get_column_letter(col)].width = widths_sales[col - 1]
ws_sales.row_dimensions[1].height = 24

for row in range(2, 201):
    bg = "FCE4D6" if row % 2 == 0 else "FFFFFF"
    for col in range(1, len(hdrs_sales) + 1):
        ws_sales.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)
ws_sales.sheet_view.showGridLines = False

# ==================== 人件費シート ====================
ws_pay = wb.create_sheet("人件費")
hdrs_pay = ["社員名", "支給総額", "控除合計", "手取り", "支払日", "備考"]
widths_pay = [15, 14, 14, 14, 12, 24]

for col, h in enumerate(hdrs_pay, 1):
    c = ws_pay.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=C_BLUE_H)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_pay.column_dimensions[get_column_letter(col)].width = widths_pay[col - 1]
ws_pay.row_dimensions[1].height = 24

employees = ["野添優", "丸田翔吾", "源地健史"]
for i, emp in enumerate(employees, 2):
    ws_pay.cell(row=i, column=1).value = emp

for row in range(2, 14):
    bg = "EBF3FB" if row % 2 == 0 else "FFFFFF"
    for col in range(1, len(hdrs_pay) + 1):
        ws_pay.cell(row=row, column=col).fill = PatternFill("solid", fgColor=bg)

# 合計行（row 15）
ws_pay.cell(row=15, column=1).value = "合計"
ws_pay.cell(row=15, column=1).font = Font(bold=True)
ws_pay.cell(row=15, column=2).value = "=SUM(B2:B14)"
ws_pay.cell(row=15, column=2).font = Font(bold=True)
ws_pay.cell(row=15, column=3).value = "=SUM(C2:C14)"
ws_pay.cell(row=15, column=3).font = Font(bold=True)
ws_pay.cell(row=15, column=4).value = "=SUM(D2:D14)"
ws_pay.cell(row=15, column=4).font = Font(bold=True)
for col in range(1, len(hdrs_pay) + 1):
    ws_pay.cell(row=15, column=col).fill = PatternFill("solid", fgColor="BDD7EE")

ws_pay.sheet_view.showGridLines = False

wb.save(filename)
print(f"✓ 生成完了: {filename}")
