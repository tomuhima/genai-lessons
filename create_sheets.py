import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 設定
EMPLOYEES = ["源地 健史", "丸田 翔吾"]
MONTHS = {
    "1月": 31, "2月": 28, "3月": 31, "4月": 30, "5月": 31, "6月": 30,
    "7月": 31, "8月": 31, "9月": 30, "10月": 31, "11月": 30, "12月": 31
}

# スタイル定義
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)
name_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
name_font = Font(bold=True, size=10)
cell_font = Font(size=10)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_header(cell, value):
    cell.value = value
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

def apply_name(cell, value):
    cell.value = value
    cell.font = name_font
    cell.fill = name_fill
    cell.alignment = left
    cell.border = border

def apply_cell(cell, value=""):
    cell.value = value
    cell.font = cell_font
    cell.alignment = center
    cell.border = border

# =============================================
# ファイル① 日報ログ.xlsx
# =============================================
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "ログ"

headers = ["氏名", "日付", "現場名", "開始時間", "終了時間", "休憩(分)", "実働時間", "作業内容"]
col_widths = [14, 14, 20, 12, 12, 12, 12, 40]

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    apply_header(ws1.cell(row=1, column=col), h)
    ws1.column_dimensions[get_column_letter(col)].width = w

ws1.row_dimensions[1].height = 22
ws1.freeze_panes = "A2"

wb1.save("/home/user/genai-lessons/日報ログ.xlsx")
print("日報ログ.xlsx 作成完了")

# =============================================
# ファイル② 月間稼働表.xlsx（年間12シート）
# =============================================
wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)  # デフォルトシート削除

for month_name, days in MONTHS.items():
    ws = wb2.create_sheet(title=month_name)

    # 1行目: ヘッダー（氏名 + 1日〜n日 + 合計）
    apply_header(ws.cell(row=1, column=1), "氏名")
    ws.column_dimensions["A"].width = 14

    for d in range(1, days + 1):
        col = d + 1
        apply_header(ws.cell(row=1, column=col), f"{d}日")
        ws.column_dimensions[get_column_letter(col)].width = 7

    # 合計列
    total_col = days + 2
    apply_header(ws.cell(row=1, column=total_col), "合計")
    ws.column_dimensions[get_column_letter(total_col)].width = 10

    ws.row_dimensions[1].height = 22

    # 2行目以降: 社員名
    for row, emp in enumerate(EMPLOYEES, 2):
        apply_name(ws.cell(row=row, column=1), emp)

        for d in range(1, days + 1):
            apply_cell(ws.cell(row=row, column=d + 1), "")

        # 合計列にSUM数式
        last_data_col = get_column_letter(days + 1)
        sum_cell = ws.cell(row=row, column=total_col)
        sum_cell.value = f"=SUM(B{row}:{last_data_col}{row})"
        sum_cell.font = Font(bold=True, size=10)
        sum_cell.alignment = center
        sum_cell.border = border

        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "B2"

wb2.save("/home/user/genai-lessons/月間稼働表.xlsx")
print("月間稼働表.xlsx 作成完了")
print("完了！")
